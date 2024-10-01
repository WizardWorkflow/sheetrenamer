import streamlit as st
from openpyxl import load_workbook
import io
import pandas as pd
import re

# Set page configuration
st.set_page_config(
    page_title="Excel Sheet Renamer",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Sidebar with instructions
st.sidebar.title("Instructions")
st.sidebar.markdown("""
**Steps to use the app:**
1. **Upload Excel Files**: Choose one or more `.xlsx` files.
2. **Enter Prefixes**: For each file, enter a prefix to add to sheet names.
3. **Edit Sheet Names (Optional)**: Manually adjust new sheet names if needed.
4. **Export Files**: Download the modified files individually or all at once.
""")

st.sidebar.markdown("---")
st.sidebar.title("Need Help?")
with st.sidebar.expander("FAQ"):
    st.markdown("""
    **Q:** What file types are supported?
    **A:** The app supports `.xlsx` files (Excel workbooks without macros).

    **Q:** How are formulas handled?
    **A:** Formulas referencing sheets within the workbook are updated to reflect new sheet names.

    **Q:** Who can I contact for support?
    **A:** Please reach out to **Ansh Gandhi** at [wizardworkflow@gmail.com](mailto:wizardworkflow@gmail.com) or call **+91 7588834433**.
    """)

st.sidebar.markdown("---")
st.sidebar.write("Developed by **Ansh Gandhi**")

# Main page title
st.title("📄 Excel Sheet Renamer")

# Initialize session state for file data
if 'file_data' not in st.session_state:
    st.session_state.file_data = []

# File upload section
st.header("1️⃣ Upload Excel Files")
uploaded_files = st.file_uploader(
    "Choose one or more Excel files to rename sheets",
    accept_multiple_files=True,
    type=['xlsx']
)

if uploaded_files:
    # Process uploaded files
    st.header("2️⃣ Enter Prefixes and Edit Sheet Names")
    for i, uploaded_file in enumerate(uploaded_files):
        file_bytes = uploaded_file.getvalue()
        file_name = uploaded_file.name

        # Load workbook to get sheet names
        try:
            wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=False)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception as e:
            st.error(f"Failed to read {file_name}: {e}")
            continue

        # Initialize session state for prefixes and new sheet names
        if f'prefix_{i}' not in st.session_state:
            st.session_state[f'prefix_{i}'] = ''
        if f'sheet_data_{i}' not in st.session_state:
            # Create a DataFrame for old and new sheet names
            df = pd.DataFrame({
                'Original Name': sheet_names,
                'New Name': sheet_names
            })
            st.session_state[f'sheet_data_{i}'] = df

        # File expander
        with st.expander(f"📝 File: {file_name}", expanded=True):
            # Prefix input with tooltip
            prefix = st.text_input(
                f"Enter prefix for {file_name}",
                value=st.session_state[f'prefix_{i}'],
                key=f"prefix_input_{i}",
                help="This prefix will be added to each sheet name."
            )
            st.session_state[f'prefix_{i}'] = prefix

            # Update new sheet names with prefix
            df = st.session_state[f'sheet_data_{i}']
            for idx in df.index:
                old_name = df.at[idx, 'Original Name']
                default_new_name = f"{prefix}.{idx+1:02d} {old_name}" if prefix else old_name
                # Ensure name is not longer than 31 chars
                default_new_name = default_new_name[:31]
                df.at[idx, 'New Name'] = default_new_name

            # Allow user to edit new sheet names
            st.markdown("#### Edit Sheet Names")
            edited_df = st.data_editor(
                df,
                num_rows="dynamic",
                key=f"data_editor_{i}",
                use_container_width=True
            )
            st.session_state[f'sheet_data_{i}'] = edited_df

            # Update file data in session state
            if len(st.session_state.file_data) <= i:
                st.session_state.file_data.append({
                    'file_name': file_name,
                    'file_bytes': file_bytes,
                    'prefix': prefix,
                    'sheet_data': edited_df
                })
            else:
                st.session_state.file_data[i]['prefix'] = prefix
                st.session_state.file_data[i]['sheet_data'] = edited_df

            # Export button for individual files
            if st.button(f"💾 Export {file_name}", key=f"export_button_{i}"):
                with st.spinner(f"Processing {file_name}..."):
                    try:
                        wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=False)
                        sheet_name_mapping = {}
                        # First, collect old and new sheet names
                        for idx, sheet in enumerate(wb.worksheets):
                            old_name = sheet.title
                            new_name = edited_df.at[idx, 'New Name']
                            # Ensure name is not longer than 31 characters
                            new_name = new_name[:31]
                            sheet_name_mapping[old_name] = new_name

                        # Update sheet names
                        for sheet in wb.worksheets:
                            sheet.title = sheet_name_mapping[sheet.title]

                        # Regular expression to find sheet references
                        sheet_names_pattern = '|'.join(re.escape(name) for name in sheet_name_mapping.keys())
                        sheet_ref_regex = re.compile(r"('?(%s)'?)!" % sheet_names_pattern)

                        # Function to replace sheet names in formulas
                        def replace_sheet_names_in_formula(formula, mapping):
                            def replacer(match):
                                sheet_ref = match.group(1)
                                sheet_name = sheet_ref.strip("'")
                                new_sheet_name = mapping.get(sheet_name, sheet_name)
                                # Add quotes if sheet name contains spaces or special characters
                                if ' ' in new_sheet_name or any(c in new_sheet_name for c in ('-', '+', '(', ')', '[', ']')):
                                    new_sheet_name = f"'{new_sheet_name}'"
                                return f"{new_sheet_name}!"
                            # Replace sheet references in the formula
                            return sheet_ref_regex.sub(replacer, formula)

                        # Update formulas in all cells
                        for sheet in wb.worksheets:
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.data_type == 'f':  # If cell contains a formula
                                        original_formula = cell.value
                                        if original_formula:
                                            new_formula = replace_sheet_names_in_formula(original_formula, sheet_name_mapping)
                                            cell.value = new_formula

                        # Save the workbook to a BytesIO object
                        output = io.BytesIO()
                        wb.save(output)
                        processed_file = output.getvalue()
                        wb.close()

                        # Provide a download button
                        st.success(f"File {file_name} processed successfully!")
                        st.download_button(
                            label=f"⬇️ Download {file_name}",
                            data=processed_file,
                            file_name=file_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download_button_{i}"
                        )
                    except Exception as e:
                        st.error(f"Failed to process the file {file_name}: {e}")

    # Export all files at once
    st.header("3️⃣ Export All Files")
    if st.button("💾 Export All Files"):
        with st.spinner("Processing all files..."):
            all_files_processed = True
            zip_buffer = io.BytesIO()
            from zipfile import ZipFile

            with ZipFile(zip_buffer, "a") as zip_file:
                for i, data in enumerate(st.session_state.file_data):
                    file_name = data['file_name']
                    file_bytes = data['file_bytes']
                    edited_df = data['sheet_data']

                    try:
                        wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=False)
                        sheet_name_mapping = {}
                        # First, collect old and new sheet names
                        for idx, sheet in enumerate(wb.worksheets):
                            old_name = sheet.title
                            new_name = edited_df.at[idx, 'New Name']
                            new_name = new_name[:31]
                            sheet_name_mapping[old_name] = new_name

                        # Update sheet names
                        for sheet in wb.worksheets:
                            sheet.title = sheet_name_mapping[sheet.title]

                        # Regular expression to find sheet references
                        sheet_names_pattern = '|'.join(re.escape(name) for name in sheet_name_mapping.keys())
                        sheet_ref_regex = re.compile(r"('?(%s)'?)!" % sheet_names_pattern)

                        # Function to replace sheet names in formulas
                        def replace_sheet_names_in_formula(formula, mapping):
                            def replacer(match):
                                sheet_ref = match.group(1)
                                sheet_name = sheet_ref.strip("'")
                                new_sheet_name = mapping.get(sheet_name, sheet_name)
                                if ' ' in new_sheet_name or any(c in new_sheet_name for c in ('-', '+', '(', ')', '[', ']')):
                                    new_sheet_name = f"'{new_sheet_name}'"
                                return f"{new_sheet_name}!"
                            return sheet_ref_regex.sub(replacer, formula)

                        # Update formulas
                        for sheet in wb.worksheets:
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.data_type == 'f':
                                        original_formula = cell.value
                                        if original_formula:
                                            new_formula = replace_sheet_names_in_formula(original_formula, sheet_name_mapping)
                                            cell.value = new_formula

                        # Save the workbook to a BytesIO object
                        output = io.BytesIO()
                        wb.save(output)
                        wb.close()

                        # Add the file to the zip
                        zip_file.writestr(file_name, output.getvalue())
                    except Exception as e:
                        st.error(f"Failed to process the file {file_name}: {e}")
                        all_files_processed = False

            if all_files_processed:
                st.success("All files processed successfully!")
                st.download_button(
                    label="⬇️ Download All Files as ZIP",
                    data=zip_buffer.getvalue(),
                    file_name="modified_excel_files.zip",
                    mime="application/zip"
                )
            else:
                st.error("Some files could not be processed.")

else:
    st.info("Please upload one or more Excel files to get started.")
